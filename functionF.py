#reading excel files
import openpyxl
import win32com.client
import pandas as pd
import sys


#global variables
NumAM = 0
NumPM = 0

#calculates the number of registrations
def numberOfReg(fileloc):
    #open workbook
    wb = openpyxl.load_workbook(fileloc)

    sheet = wb.get_sheet_by_name('Class List')

    sheet['A1'].value

    sh = wb.active
    numReg = 0

    for i in range(7, sh.max_row+1):
        cell = 'A' + str(i)
        if isinstance(sheet[cell].value, int) == True :
            numReg += 1
            #print(sheet[cell].value)  
    print("TOTAL NUMBER OF REGISTRATIONS: ", numReg)

#calculates all the half day registrations and prints it out
def half(fileloc):
    wb = openpyxl.load_workbook(fileloc)

    sheet = wb.get_sheet_by_name('Class List')


    sh = wb.active
    halfam1 = 0
    halfam2 = 0
    halfam3 = 0
    halfam4 = 0
    halfam5 = 0
    halfam6 = 0

    halfpm1 = 0
    halfpm2 = 0
    halfpm3 = 0
    halfpm4 = 0
    halfpm5 = 0
    halfpm6 = 0

    for i in range(7, sh.max_row+1):
        cell1 = 'I' + str(i)
        cell2 = 'H' + str(i)
        
        if sheet[cell1].value == "Bike Half (main price): 09:00 AM - 12:00 PM":
            if sheet[cell2].value == "Level 1 - Newbees":
                halfam1+= 1
            if sheet[cell2].value == "Level 2 - Advanced Newbees":
                halfam2 += 1
            if sheet[cell2].value == "Level 3 - Pedalheads":
                halfam3 += 1
            if sheet[cell2].value == "Level 4 - Advanced Pedalheads":
                halfam4 += 1
            if sheet[cell2].value == "Level 5 - Gearheads":
                halfam5 += 1
            if sheet[cell2].value == "Level 6 - Treadheads":
                halfam6 += 1
        if sheet[cell1].value == "Bike Half (main price): 01:00 PM - 04:00 PM":
            if sheet[cell2].value == "Level 1 - Newbees":
                halfpm1+= 1
            if sheet[cell2].value == "Level 2 - Advanced Newbees":
                halfpm2 += 1
            if sheet[cell2].value == "Level 3 - Pedalheads":
                halfpm3 += 1
            if sheet[cell2].value == "Level 4 - Advanced Pedalheads":
                halfpm4 += 1
            if sheet[cell2].value == "Level 5 - Gearheads":
                halfpm5 += 1
            if sheet[cell2].value == "Level 6 - Treadheads":
                halfpm6 += 1
    print("========================================\n========================================")
    print("NUMBER OF HALF DAY AM LEVEL 1:", halfam1, "registrations")
    print("NUMBER OF HALF DAY AM LEVEL 2:", halfam2, "registrations")
    print("NUMBER OF HALF DAY AM LEVEL 3:", halfam3, "registrations")
    print("NUMBER OF HALF DAY AM LEVEL 4:", halfam4, "registrations")
    print("NUMBER OF HALF DAY AM LEVEL 5:", halfam5, "registrations")
    print("NUMBER OF HALF DAY AM LEVEL 6:", halfam6, "registrations")
    print("========================================\n========================================")
    print("NUMBER OF HALF DAY PM LEVEL 1:", halfpm1, "registrations")
    print("NUMBER OF HALF DAY PM LEVEL 2:", halfpm2, "registrations")
    print("NUMBER OF HALF DAY PM LEVEL 3:", halfpm3, "registrations")
    print("NUMBER OF HALF DAY PM LEVEL 4:", halfpm4, "registrations")
    print("NUMBER OF HALF DAY PM LEVEL 5:", halfpm5, "registrations")
    print("NUMBER OF HALF DAY PM LEVEL 6:", halfpm6, "registrations")

#calculates all the allday registrations and prints it out
def allday(fileloc):
    wb = openpyxl.load_workbook(fileloc)

    sheet = wb.get_sheet_by_name('Class List')

    sh = wb.active

    all1 = 0
    all2 = 0
    all3 = 0
    all4 = 0
    all5 = 0
    all6 = 0

    for i in range(7, sh.max_row+1):
        cell1 = 'I' + str(i)
        cell2 = 'H' + str(i)

        if sheet[cell1].value == "Bike All (main price): 09:00 AM - 04:00 PM":
            if sheet[cell2].value == "Level 1 - Newbees":
                all1+= 1
            if sheet[cell2].value == "Level 2 - Advanced Newbees":
                all2 += 1
            if sheet[cell2].value == "Level 3 - Pedalheads":
                all3 += 1
            if sheet[cell2].value == "Level 4 - Advanced Pedalheads":
                all4 += 1
            if sheet[cell2].value == "Level 5 - Gearheads":
                all5 += 1
            if sheet[cell2].value == "Level 6 - Treadheads":
                all6 += 1
    print("========================================\n========================================")
    print("NUMBER OF ALL DAY LEVEL 1:", all1, "registrations")
    print("NUMBER OF ALL DAY LEVEL 2:", all2, "registrations")
    print("NUMBER OF ALL DAY LEVEL 3:", all3, "registrations")
    print("NUMBER OF ALL DAY LEVEL 4:", all4, "registrations")
    print("NUMBER OF ALL DAY LEVEL 5:", all5, "registrations")
    print("NUMBER OF ALL DAY LEVEL 6:", all6, "registrations")


#copies all the half day am and all day in a file
def copyRow(fileloc,filepath):
    #opening workbook
    wb = openpyxl.load_workbook(fileloc)

    #opening proper sheet
    sheet = wb.get_sheet_by_name('Class List')

    #setting variables
    sh = wb.active

    #setting variables for the new excel sheet
    wb2 = openpyxl.Workbook()

    ws2 = wb2.active
    ws2.title = "Class List"

    ws3 = wb2.create_sheet(0)
    ws3.title = "AM & AD"

    ws4 = wb2.create_sheet(0)
    ws4.title = "PM & AD"

    ws5 = wb2.create_sheet(0)
    ws5.title = "ALPHA AM & AD"

    ws6 = wb2.create_sheet(0)
    ws6.title = "ALPHA PM & AD"

    #max row and max column
    mr = sh.max_row
    mc = sh.max_column 
    for k in range (1, 6):
        for j in range (1, mc + 1):
            c = sh.cell(row = k, column = j)
            ws2.cell(row = k, column = j).value = c.value
            ws3.cell(row = k, column = j).value = c.value
            ws4.cell(row = k, column = j).value = c.value
        for alphaC in range (2, mc + 1):
            alpha = sh.cell(row = k, column = alphaC)
            ws5.cell(row = k, column = alphaC-1).value = alpha.value
            ws6.cell(row = k, column = alphaC-1).value = alpha.value

    rowTrackingAM = 6
    rowTrackingPM = 6
    for i in range(7, mr+1):
        orderNum = 'A' + str(i)
        regTime = 'I' + str(i)

        #COPYING ORIGINAL SET INTO FILE
        for OG in range (1, mc + 1):
            c = sh.cell(row = i, column = OG)
            ws2.cell(row = i, column = OG).value = c.value

        #CHECKING IF THERE IS A REGISTRATION
        if isinstance(sheet[orderNum].value, int) == True :
        #COPYING FOR AM AND AD
            if sheet[regTime].value == "Bike All (main price): 09:00 AM - 04:00 PM" or sheet[regTime].value == "Bike Half (main price): 09:00 AM - 12:00 PM": 
                rowTrackingAM += 1
                for j in range (1, mc + 1):
                    #for x in range(7, mr+7):
                    c = sh.cell(row = i, column = j)
                    ws3.cell(row = rowTrackingAM, column = j).value = c.value
        #ALPHA AM AND AD
                for m in range (2, mc + 1):
                    #for x in range(7, mr+7):
                    c = sh.cell(row = i, column = m)
                    ws5.cell(row = rowTrackingAM, column = m-1).value = c.value
        #COPYING FOR PM AND AD
            if sheet[regTime].value == "Bike All (main price): 09:00 AM - 04:00 PM" or sheet[regTime].value == "Bike Half (main price): 01:00 PM - 04:00 PM":
                rowTrackingPM += 1 
                for k in range (1, mc + 1):
                    #for x in range(7, mr+7):
                    c = sh.cell(row = i, column = k)
                    ws4.cell(row = rowTrackingPM, column = k).value = c.value
        #ALPHA PM AND AD
                for l in range (2, mc + 1):
                    #for x in range(7, mr+7):
                    c = sh.cell(row = i, column = l)
                    ws6.cell(row = rowTrackingPM, column = l-1).value = c.value
    #putting alpha sheets in alphabetical order

    rangeALPHAAM = 'A6:A' + str(ws5.max_row)
    
    #ws5.Range(rangeALPHAAM).Sort(Key1=ws5.Range('A5'), Order1=1, Orientation=1)

    #creating a new file
    wb2.save(filepath)
    NumAM = ws5.max_row

#this function doesn't work :( im not sure why, 
#it's probably cause it can't edit a file that is being created from openpxyl? I'm not sure
def orderSorting(filepath):
    #not sure what this does
    excel = win32com.client.Dispatch("Excel.Application")
    print("this works")

    #opening the excel file that was JUST created from the previous function
    wb = excel.Workbooks.Open(filepath)
    ws = wb.Worksheets('ALPHA AM & AD')
    
    #creating other variables
    xlAscending = 1
    xlSortColumns = 0
    xlYes = 1



    rangeALPHAAM = 'A6:A' + str(NumAM)
    #LastRow = Sheets("ALPHA AM & AD").Range("A" & Sheets("ALPHA AM & AD").Rows.Count).End(xlUp).Row

    ws.Range(rangeALPHAAM).Sort(Key1=ws.Range('A5'), Order1=xlAscending,
                                Key2=ws.Range('B5'), Order2=xlAscending, 
                                Key3=ws.Range('C5'), Order3=xlAscending,
                                header=xlYes, Orientation=xlSortColumns)

    wb.Save()
    excel.Application.Quit
